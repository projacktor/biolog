import math
import openpyxl as xl
import os

import table_headers as headers
from functions import *


class Calculator:
    @staticmethod
    def fill_the_table(name: str):
        wb: xl.Workbook = xl.Workbook()
        ws = wb[wb.sheetnames[0]]
        ws.column_dimensions['A'].width = 40
        for i in range(2, len(headers.headers) + 2):
            ws["A" + str(i)] = headers.headers[i - 2]

        for i in range(36, len(headers.headers_res) + 36):
            ws["A" + str(i)] = headers.headers_res[i - 36]

        for i in range(46, len(headers.indexes) + 46):
            ws["A" + str(i)] = headers.indexes[i - 46]
        wb.save(f"output{name}.xlsx")
        return f"output{name}.xlsx"

    @staticmethod
    def calculate_logic(input_path: str, output_path: str, column: int):
        # open files
        output_file: xl.Workbook = xl.load_workbook(output_path)
        input_file: xl.Workbook = xl.load_workbook(input_path)
        output_sheet = output_file[output_file.sheetnames[0]]
        input_sheet = input_file[input_file.sheetnames[0]]

        # fill plates header
        plate_name: str = str(input_sheet.cell(row=4, column=1).value)
        plate_name = plate_name[plate_name.find(":") + 1:].lstrip()
        output_sheet.cell(row=1, column=column, value=plate_name)

        # fill the dictionary with headers
        substrates: dict = dict()
        for i in range(2, len(headers.headers) + 2):
            substrates[headers.headers[i - 2]] = 0

        # counting average water
        water: float = arithmetical_avg(
            [float(input_sheet.cell(row=7, column=2).value),
             float(input_sheet.cell(row=7, column=6).value),
             float(input_sheet.cell(row=7, column=10).value)])

        # counting average of substrates - @water
        items: list = list(substrates.items())
        index: int = 0
        all_values_sum: float = 0
        for i in range(7, 14 + 1):
            for j in range(2, 5 + 1):
                if i == 7 and j == 2:
                    continue
                key_to_change: str = items[index][0]
                el1 = float(input_sheet.cell(row=i, column=j).value)
                el2 = float(input_sheet.cell(row=i, column=j + 4).value)
                el3 = float(input_sheet.cell(row=i, column=j + 8).value)
                if arithmetical_avg([el1, el2, el3]) - water >= 0:
                    substrates[key_to_change] = arithmetical_avg([el1, el2, el3]) - water
                else:
                    substrates[key_to_change] = 0

                all_values_sum += substrates[key_to_change]
                index += 1

        # print to table
        for i in range(30 + 1):
            output_sheet.cell(row=i + 2, column=column, value=substrates[items[i][0]])
            # output_sheet["B" + str(i + 2)] = substrates[items[i][0]]

        # SWCD counting
        amino_group: float = arithmetical_avg([substrates[_] for _ in headers.amino_group])
        amino_acids: float = arithmetical_avg([substrates[_] for _ in headers.amino_acids])
        polymers: float = arithmetical_avg([substrates[_] for _ in headers.polymers])
        phenolic: float = arithmetical_avg([substrates[_] for _ in headers.phenolic])
        carbonic_acids: float = arithmetical_avg([substrates[_] for _ in headers.carbonic_acids])
        carbohydrates: float = arithmetical_avg([substrates[_] for _ in headers.carbohydrates])

        swcd: list = [amino_group, phenolic, polymers, amino_acids, carbonic_acids, carbohydrates]

        # SWCD output
        for i in range(37, 42 + 1):
            output_sheet.cell(row=i, column=column, value=swcd[i - 37])
            # output_sheet["B" + str(i)] = swcd[i - 37]

        # ACWD output
        output_sheet.cell(row=46, column=column, value=float(all_values_sum / 31))
        # output_sheet["B46"] = float(all_values_sum / 31)

        # Shenon index counting
        p: dict = {key: value / all_values_sum for key, value in substrates.items()}

        log: dict = dict()
        for k, v in p.items():
            if v == 0.0:
                log[k] = 0
            else:
                log[k] = math.log(v)

        p_l: dict = dict()
        for k, v in p.items():
            p_l[k] = v * log[k]

        shenon_index: float = sum(p_l.values()) * -1

        # Shenon index output
        output_sheet.cell(row=47, column=column, value=shenon_index)
        # output_sheet["B47"] = shenon_index

        output_file.save(output_path)

    def calculate_files_for_one(self, files: list):
        output_path: str = self.fill_the_table("")
        column: int = 2
        for file in files:
            flag: bool = False
            if file.endswith(".xls") and not(file.endswith(".xlsx")):
                file = convert_xls_to_xlsx(file)
                flag = True
            self.calculate_logic(file, output_path, column)
            if flag:
                os.remove(file)
            column += 1

    def calculate_files_for_many(self, files: list):
        for file in files:
            flag: bool = False
            if file.endswith(".xls") and not(file.endswith(".xlsx")):
                file = convert_xls_to_xlsx(file)
                flag = True
            output_file: str = ""
            for i in file.split("\\"):
                if i.endswith(".xlsx"):
                    output_file: str = self.fill_the_table("_" + i.rstrip(".xlsx"))
                    break
            self.calculate_logic(file, output_file, 2)
            if flag:
                os.remove(file)
