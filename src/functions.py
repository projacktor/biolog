import xlrd as xd
from openpyxl import Workbook


def arithmetical_avg(numbers: list) -> float:
    return sum(numbers) / len(numbers)


def convert_xls_to_xlsx(path: str) -> str:
    xls: xd.Workbook = xd.open_workbook(path)
    xlsx: Workbook = Workbook()
    default_sheet = xlsx.active
    xlsx.remove(default_sheet)
    for sheet_index in range(xls.nsheets):
        xls_sheet = xls.sheet_by_index(sheet_index)
        xlsx_sheet = xlsx.create_sheet(title=xls_sheet.name)

        for row in range(xls_sheet.nrows):
            for col in range(xls_sheet.ncols):
                xlsx_sheet.cell(row=row + 1, column=col + 1).value = xls_sheet.cell_value(row, col)

    new_path: str = path.replace(".xls", ".xlsx")
    xlsx.save(new_path)
    return new_path
